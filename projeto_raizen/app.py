#!/usr/bin/env python3
# -*- encoding: utf-8 -*-

import pandas as pd #join to lib xlrd
from openpyxl import load_workbook
from datetime import datetime
import csv

class get_data_workbook:

    def __init__(self):
        self._raw_data_path: str = "C:\\Users\\ramon\Desktop\\projeto_raizen\\data\\temp_data\\vendas-combustiveis-m3.xls"
        self._target_data_path: str = "C:\\Users\\ramon\\Desktop\\projeto_raizen\\data\\target_data\\vendas-combustiveis-m3.xlsx"
        self._csv_derivative_path: str = r"\\wsl.localhost\\Ubuntu-18.04\\home\\ramon\\projeto_raizen\\data\\csv_data\\oil_derivative.csv"
        self._csv_diesel_path: str = r"\\wsl.localhost\\Ubuntu-18.04\\home\\ramon\\projeto_raizen\\data\\csv_data\\diesel.csv"
        self._read_spreadsheet = None
        self._sheet_obj = None
        self._data_list_1: list = list()
        self._data_list_2: list = list()


    def convert_xls_to_xlsx(self):
        raw_data = pd.read_excel(self._raw_data_path)
        raw_data.to_excel(self._target_data_path)


    def read_and_set_spreadsheet(self):
        self._read_spreadsheet = load_workbook(self._target_data_path)
        self._sheet_obj = self._read_spreadsheet.active


    def get_date_and_value_1(self, coluna: int, year: str, inicio: int, fim: int) -> list:
        month = 1
        for linha in range(inicio, fim + 1):
            cell_data = self._sheet_obj.cell(row=linha, column=coluna).value
            self._data_list_1.insert(0, ["%s-%s" % (year, month), "AMAZONAS", "ETANOL_HIDRATADO", "METROS_CUBICOS",  cell_data, datetime.now()])
            month += 1


    def get_data_oil_derivative_uf_product(self):
        col = 4 # 4 - 22
        for ano in range(2000, 2021):
            self.get_date_and_value_1(col, ano, 54, 65)
            col += 1


    def get_date_and_value_2(self, coluna: int, year: str, inicio: int, fim: int) -> list:
        month = 1
        for linha in range(inicio, fim + 1):
            cell_data = self._sheet_obj.cell(row=linha, column=coluna).value
            self._data_list_2.insert(0, ["%s-%s" % (year, month), "AMAZONAS", "OLEO_DIESEL_MARITIMO", "METROS_CUBICOS",  cell_data, datetime.now()])
            month += 1


    def get_data_diesel_uf_type(self):
        col = 4 # 4 - 11
        for ano in range(2013, 2021):
            self.get_date_and_value_2(col, ano, 134, 145)
            col += 1


    def save_data_csv_derivative(self):
        with open(self._csv_derivative_path, "w") as derivative:
            derivative_csv = csv.writer(derivative)
            derivative_csv.writerows(self._data_list_1)


    def save_data_csv_diesel(self):
        with open(self._csv_diesel_path, "w") as diesel:
            diesel_csv = csv.writer(diesel)
            diesel_csv.writerows(self._data_list_2)



if(__name__ != '__main__'):
    pass
else:
    obj1 = get_data_workbook()
    obj1.convert_xls_to_xlsx()
    obj1.read_and_set_spreadsheet()
    obj1.get_data_oil_derivative_uf_product()
    obj1.get_data_diesel_uf_type()
    obj1.save_data_csv_derivative()
    obj1.save_data_csv_diesel()
    print("RC = 0")
    #print(obj1._data_list_1)
    #print(obj1._data_list_2)